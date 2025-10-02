"""Servicio de sincronización de datos desde Google Drive hacia MySQL.

Emula el comportamiento de ``gscripts/extraer_datos.js`` utilizando Python,
aplicando patrones de mapeo modernos, validaciones y logging estructurado.
"""

from __future__ import annotations

import io
import json
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from math import isnan
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence

import structlog
from openpyxl import load_workbook
from pydantic import BaseModel, Field
from sqlalchemy import (
    Boolean,
    Column,
    Date,
    DateTime,
    Integer,
    MetaData,
    String,
    Table,
    create_engine,
    insert,
    select,
)
from sqlalchemy.engine import Engine
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.orm import Session, sessionmaker

from app.config import PROJECT_ROOT, Settings
from app.services.drive_service import DriveService


EXCEL_MIME_TYPES = [
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel",
]


def _is_null(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and not value.strip():
        return True
    if isinstance(value, float) and isnan(value):
        return True
    return False


def _normalize_boolean(value: Any, default: bool = False) -> bool:
    if _is_null(value):
        return default
    if isinstance(value, bool):
        return value
    str_value = str(value).strip().lower()
    return str_value in {"1", "true", "verdadero", "sí", "si", "x", "yes"}


def _normalize_date(value: Any) -> Optional[str]:
    if _is_null(value):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float)) and not isnan(value):
        # Excel serial date handling (assuming 1899-12-30 origin)
        base = datetime(1899, 12, 30)
        try:
            return (base + timedelta(days=float(value))).date().isoformat()
        except Exception:  # noqa: BLE001
            return None
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None
        date_formats = [
            "%Y-%m-%d",
            "%d/%m/%Y",
            "%Y/%m/%d",
            "%d-%m-%Y",
            "%m/%d/%Y",
            "%d-%b-%Y",
            "%d-%b-%y",
            "%d %b %Y",
            "%Y%m%d",
        ]
        for fmt in date_formats:
            try:
                return datetime.strptime(value, fmt).date().isoformat()
            except ValueError:
                continue
    return None


def _normalize_fecha_presencial(value: Any) -> Optional[str]:
    if _is_null(value):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float)) and not isnan(value):
        base = datetime(1899, 12, 30)
        try:
            return (base + timedelta(days=float(value))).date().isoformat()
        except Exception:  # noqa: BLE001
            return None
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None
        date_formats = [
            "%d/%m/%Y",
            "%d-%m-%Y",
            "%Y-%m-%d",
            "%m/%d/%Y",
        ]
        for fmt in date_formats:
            try:
                return datetime.strptime(value, fmt).date().isoformat()
            except ValueError:
                continue
    return None


class MissionaryRecord(BaseModel):
    """DTO principal que representa una fila lista para insertarse en MySQL."""

    id: int
    id_distrito: Optional[str] = None
    tipo: Optional[str] = None
    rama: Optional[int] = None
    distrito: Optional[str] = None
    pais: Optional[str] = None
    numero_lista: Optional[int] = None
    numero_companerismo: Optional[int] = None
    tratamiento: Optional[str] = None
    nombre_misionero: str
    companero: Optional[str] = None
    mision_asignada: Optional[str] = None
    estaca: Optional[str] = None
    hospedaje: Optional[str] = None
    foto: Optional[str] = None
    fecha_llegada: Optional[str] = None
    fecha_salida: Optional[str] = None
    fecha_generacion: Optional[str] = None
    comentarios: Optional[str] = None
    investido: bool = False
    fecha_nacimiento: Optional[str] = None
    foto_tomada: bool = False
    pasaporte: bool = False
    folio_pasaporte: Optional[str] = None
    fm: Optional[str] = None
    ipad: bool = False
    closet: Optional[str] = None
    llegada_secundaria: Optional[str] = None
    pday: Optional[str] = None
    host: bool = False
    tres_semanas: bool = False
    device: bool = False
    correo_misional: Optional[str] = None
    correo_personal: Optional[str] = None
    fecha_presencial: Optional[str] = None
    activo: bool = True

    class Config:
        arbitrary_types_allowed = True

    @classmethod
    def from_row(cls, row: Sequence[Any], *, logger: structlog.stdlib.BoundLogger, excel_file_id: str, row_index: int) -> Optional["MissionaryRecord"]:
        try:
            missionary_id_raw = row[0]
        except IndexError:
            logger.warning(
                "⚠️ Fila sin datos suficientes",
                etapa="normalizacion",
                excel_file_id=excel_file_id,
                row_index=row_index,
                table_errors=["row_without_id"],
                message_id=None,
            )
            return None

        if missionary_id_raw in (None, "", float("nan")):
            logger.warning(
                "⚠️ Fila ignorada por carecer de ID",
                etapa="normalizacion",
                excel_file_id=excel_file_id,
                row_index=row_index,
                table_errors=["missing_id"],
                message_id=None,
            )
            return None

        try:
            missionary_id = int(str(missionary_id_raw).strip())
        except (TypeError, ValueError):
            logger.warning(
                "⚠️ ID inválido en fila",
                etapa="normalizacion",
                excel_file_id=excel_file_id,
                row_index=row_index,
                table_errors=["invalid_id"],
                message_id=None,
            )
            return None

        def safe_int(value: Any) -> Optional[int]:
            if value in (None, "", float("nan")):
                return None
            try:
                return int(str(value).strip())
            except (TypeError, ValueError):
                return None

        registro = cls(
            id=missionary_id,
            id_distrito=str(row[1]).strip() if len(row) > 1 and row[1] not in (None, "", float("nan")) else None,
            tipo=str(row[2]).strip() if len(row) > 2 and row[2] not in (None, "", float("nan")) else None,
            rama=safe_int(row[3]) if len(row) > 3 else None,
            distrito=str(row[4]).strip() if len(row) > 4 and row[4] not in (None, "", float("nan")) else None,
            pais=str(row[5]).strip().title() if len(row) > 5 and row[5] not in (None, "", float("nan")) else None,
            numero_lista=safe_int(row[6]) if len(row) > 6 else None,
            numero_companerismo=safe_int(row[7]) if len(row) > 7 else None,
            tratamiento=None,
            nombre_misionero=str(row[8]).strip() if len(row) > 8 and row[8] not in (None, "", float("nan")) else "",
            companero=str(row[9]).strip() if len(row) > 9 and row[9] not in (None, "", float("nan")) else None,
            mision_asignada=str(row[10]).strip() if len(row) > 10 and row[10] not in (None, "", float("nan")) else None,
            estaca=str(row[11]).strip() if len(row) > 11 and row[11] not in (None, "", float("nan")) else None,
            hospedaje=str(row[12]).strip() if len(row) > 12 and row[12] not in (None, "", float("nan")) else None,
            foto=str(row[13]).strip() if len(row) > 13 and row[13] not in (None, "", float("nan")) else None,
            fecha_llegada=_normalize_date(row[14]) if len(row) > 14 else None,
            fecha_salida=_normalize_date(row[15]) if len(row) > 15 else None,
            fecha_generacion=_normalize_date(row[16]) if len(row) > 16 else None,
            comentarios=str(row[17]).strip() if len(row) > 17 and row[17] not in (None, "", float("nan")) else None,
            investido=_normalize_boolean(row[18]) if len(row) > 18 else False,
            fecha_nacimiento=_normalize_date(row[19]) if len(row) > 19 else None,
            foto_tomada=_normalize_boolean(row[20]) if len(row) > 20 else False,
            pasaporte=_normalize_boolean(row[21]) if len(row) > 21 else False,
            folio_pasaporte=str(row[22]).strip() if len(row) > 22 and row[22] not in (None, "", float("nan")) else None,
            fm=str(row[23]).strip() if len(row) > 23 and row[23] not in (None, "", float("nan")) else None,
            ipad=_normalize_boolean(row[24]) if len(row) > 24 else False,
            closet=str(row[25]).strip() if len(row) > 25 and row[25] not in (None, "", float("nan")) else None,
            llegada_secundaria=str(row[26]).strip() if len(row) > 26 and row[26] not in (None, "", float("nan")) else None,
            pday=str(row[27]).strip() if len(row) > 27 and row[27] not in (None, "", float("nan")) else None,
            host=_normalize_boolean(row[28]) if len(row) > 28 else False,
            tres_semanas=_normalize_boolean(row[29]) if len(row) > 29 else False,
            device=_normalize_boolean(row[30]) if len(row) > 30 else False,
            correo_misional=str(row[31]).strip() if len(row) > 31 and row[31] not in (None, "", float("nan")) else None,
            correo_personal=str(row[32]).strip() if len(row) > 32 and row[32] not in (None, "", float("nan")) else None,
            fecha_presencial=_normalize_fecha_presencial(row[33]) if len(row) > 33 else None,
        )

        if not registro.nombre_misionero:
            logger.warning(
                "⚠️ Registro sin nombre, se continuará",
                etapa="normalizacion",
                excel_file_id=excel_file_id,
                row_index=row_index,
                table_errors=["missing_nombre_misionero"],
                message_id=None,
            )
        return registro

    def to_database_payload(self, *, timestamp: datetime) -> Dict[str, Any]:
        return {
            "id": self.id,
            "id_distrito": self.id_distrito,
            "tipo": self.tipo,
            "rama": self.rama,
            "distrito": self.distrito,
            "pais": self.pais,
            "numero_lista": self.numero_lista,
            "numero_companerismo": self.numero_companerismo,
            "tratamiento": self.tratamiento,
            "nombre_misionero": self.nombre_misionero,
            "companero": self.companero,
            "mision_asignada": self.mision_asignada,
            "estaca": self.estaca,
            "hospedaje": self.hospedaje,
            "foto": self.foto,
            "fecha_llegada": self.fecha_llegada,
            "fecha_salida": self.fecha_salida,
            "fecha_generacion": self.fecha_generacion,
            "comentarios": self.comentarios,
            "investido": self.investido,
            "fecha_nacimiento": self.fecha_nacimiento,
            "foto_tomada": self.foto_tomada,
            "pasaporte": self.pasaporte,
            "folio_pasaporte": self.folio_pasaporte,
            "fm": self.fm,
            "ipad": self.ipad,
            "closet": self.closet,
            "llegada_secundaria": self.llegada_secundaria,
            "pday": self.pday,
            "host": self.host,
            "tres_semanas": self.tres_semanas,
            "device": self.device,
            "correo_misional": self.correo_misional,
            "correo_personal": self.correo_personal,
            "fecha_presencial": self.fecha_presencial,
            "activo": self.activo,
            "created_at": timestamp,
            "updated_at": timestamp,
        }


@dataclass
class DatabaseSyncReport:
    fecha_generacion: str
    drive_folder_id: str
    processed_files: List[Dict[str, Any]] = field(default_factory=list)
    inserted_count: int = 0
    skipped_count: int = 0
    errors: List[Dict[str, Any]] = field(default_factory=list)
    duration_seconds: float = 0.0
    continuation_token: Optional[str] = None

    def to_dict(self) -> Dict[str, Any]:
        return {
            "fecha_generacion": self.fecha_generacion,
            "drive_folder_id": self.drive_folder_id,
            "processed_files": self.processed_files,
            "inserted_count": self.inserted_count,
            "skipped_count": self.skipped_count,
            "errors": self.errors,
            "duration_seconds": self.duration_seconds,
            "continuation_token": self.continuation_token,
        }


@dataclass
class DatabaseSyncState:
    last_processed_file_id: Optional[str] = None
    continuation_token: Optional[str] = None


class DatabaseSyncStateRepository:
    """Persistencia ligera del estado de sincronización."""

    def __init__(self, path: Path):
        self.path = path
        self.path.parent.mkdir(parents=True, exist_ok=True)

    def load(self, folder_id: str) -> DatabaseSyncState:
        if not self.path.exists():
            return DatabaseSyncState()
        try:
            data = json.loads(self.path.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            return DatabaseSyncState()

        folder_state = data.get("folders", {}).get(folder_id, {})
        return DatabaseSyncState(
            last_processed_file_id=folder_state.get("last_processed_file_id"),
            continuation_token=folder_state.get("continuation_token"),
        )

    def _write_state(self, *, folder_id: str, state: DatabaseSyncState) -> None:
        if self.path.exists():
            try:
                data = json.loads(self.path.read_text(encoding="utf-8"))
            except json.JSONDecodeError:
                data = {}
        else:
            data = {}

        folders = data.setdefault("folders", {})
        folders[folder_id] = {
            "last_processed_file_id": state.last_processed_file_id,
            "continuation_token": state.continuation_token,
            "updated_at": datetime.utcnow().isoformat(),
        }
        self.path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

    def mark_processed(self, folder_id: str, file_id: str) -> None:
        self._write_state(folder_id=folder_id, state=DatabaseSyncState(last_processed_file_id=file_id))

    def mark_interrupted(self, folder_id: str, file_id: str) -> None:
        self._write_state(
            folder_id=folder_id,
            state=DatabaseSyncState(last_processed_file_id=None, continuation_token=file_id),
        )

    def clear(self, folder_id: str) -> None:
        if not self.path.exists():
            return
        try:
            data = json.loads(self.path.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            self.path.unlink(missing_ok=True)
            return

        folders = data.get("folders", {})
        if folder_id in folders:
            folders.pop(folder_id)
            if folders:
                data["folders"] = folders
                self.path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
            else:
                self.path.unlink(missing_ok=True)


metadata = MetaData()

ccm_generaciones_table = Table(
    "ccm_generaciones",
    metadata,
    Column("id", Integer, primary_key=True),
    Column("id_distrito", String(255)),
    Column("tipo", String(255)),
    Column("rama", Integer),
    Column("distrito", String(255)),
    Column("pais", String(255)),
    Column("numero_lista", Integer),
    Column("numero_companerismo", Integer),
    Column("tratamiento", String(255)),
    Column("nombre_misionero", String(255), nullable=False),
    Column("companero", String(255)),
    Column("mision_asignada", String(255)),
    Column("estaca", String(255)),
    Column("hospedaje", String(255)),
    Column("foto", String(255)),
    Column("fecha_llegada", Date),
    Column("fecha_salida", Date),
    Column("fecha_generacion", Date),
    Column("comentarios", String(1024)),
    Column("investido", Boolean, default=False),
    Column("fecha_nacimiento", Date),
    Column("foto_tomada", Boolean, default=False),
    Column("pasaporte", Boolean, default=False),
    Column("folio_pasaporte", String(255)),
    Column("fm", String(255)),
    Column("ipad", Boolean, default=False),
    Column("closet", String(255)),
    Column("llegada_secundaria", String(255)),
    Column("pday", String(255)),
    Column("host", Boolean, default=False),
    Column("tres_semanas", Boolean, default=False),
    Column("device", Boolean, default=False),
    Column("correo_misional", String(255)),
    Column("correo_personal", String(255)),
    Column("fecha_presencial", Date),
    Column("activo", Boolean, default=True),
    Column("created_at", DateTime, nullable=False),
    Column("updated_at", DateTime, nullable=False),
)


class DatabaseSyncService:
    """Orquesta la sincronización completa de una generación."""

    def __init__(
        self,
        settings: Settings,
        drive_service: DriveService,
        *,
        engine: Optional[Engine] = None,
        state_repository: Optional[DatabaseSyncStateRepository] = None,
    ):
        self.settings = settings
        self.drive_service = drive_service
        self.logger = structlog.get_logger("database_sync").bind(
            servicio="database_sync",
        )
        self.engine = engine or self._create_engine(settings)
        self.session_factory = sessionmaker(bind=self.engine, expire_on_commit=False)
        state_path = PROJECT_ROOT / "data" / "state" / "database_sync_state.json"
        self.state_repository = state_repository or DatabaseSyncStateRepository(state_path)

    def _create_engine(self, settings: Settings) -> Engine:
        url = settings.database_url
        if url:
            if url.startswith("mysql://"):
                url = url.replace("mysql://", "mysql+pymysql://", 1)
        else:
            url = (
                f"mysql+pymysql://{settings.db_user}:{settings.db_password}"
                f"@{settings.db_host}:{settings.db_port}/{settings.db_name}"
            )
        engine = create_engine(url, pool_pre_ping=True, pool_recycle=3600)
        return engine

    def sync_generation(
        self,
        fecha_generacion: str,
        drive_folder_id: str,
        *,
        force: bool = False,
    ) -> DatabaseSyncReport:
        start_time = datetime.utcnow()
        report = DatabaseSyncReport(
            fecha_generacion=fecha_generacion,
            drive_folder_id=drive_folder_id,
        )

        self.logger.info(
            "🗂️ Iniciando sincronización de generación",
            etapa="inicio",
            fecha_generacion=fecha_generacion,
            drive_folder_id=drive_folder_id,
            force=force,
            message_id=None,
        )

        state = self.state_repository.load(drive_folder_id)
        skip_until = None if force else state.last_processed_file_id
        continuation_override = None if force else state.continuation_token

        try:
            files = self.drive_service.list_folder_files(
                drive_folder_id,
                mime_types=EXCEL_MIME_TYPES,
            )
        except Exception as exc:  # noqa: BLE001
            self.logger.error(
                "❌ Error listando archivos en Drive",
                etapa="drive_list",
                drive_folder_id=drive_folder_id,
                error=str(exc),
                error_code="drive_listing_failed",
                table_errors=[],
                message_id=None,
            )
            report.errors.append(
                {
                    "code": "drive_listing_failed",
                    "message": str(exc),
                }
            )
            report.duration_seconds = (datetime.utcnow() - start_time).total_seconds()
            report.continuation_token = continuation_override or skip_until
            return report

        files_sorted = sorted(files, key=lambda item: item.get("modifiedTime", item.get("name", "")))

        skip_mode = bool(skip_until)
        if continuation_override:
            skip_mode = False

        inserted_total = 0
        skipped_total = 0

        for file_meta in files_sorted:
            file_id = file_meta.get("id")
            if not file_id:
                continue

            if continuation_override and file_id != continuation_override:
                continue
            if continuation_override and file_id == continuation_override:
                continuation_override = None  # re-procesar el archivo fallido

            if skip_mode:
                if file_id == skip_until:
                    skip_mode = False
                continue

            filename = file_meta.get("name", "archivo.xlsx")
            self.logger.info(
                "📥 Descargando archivo de generación",
                etapa="descarga_excel",
                drive_folder_id=drive_folder_id,
                excel_file_id=file_id,
                excel_filename=filename,
                message_id=None,
            )

            try:
                file_bytes = self.drive_service.download_file(file_id)
            except Exception as exc:  # noqa: BLE001
                self.logger.error(
                    "❌ Error descargando archivo desde Drive",
                    etapa="descarga_excel",
                    drive_folder_id=drive_folder_id,
                    excel_file_id=file_id,
                    excel_filename=filename,
                    error=str(exc),
                    error_code="drive_download_failed",
                    table_errors=[],
                    message_id=None,
                )
                report.errors.append(
                    {
                        "code": "drive_download_failed",
                        "message": str(exc),
                        "file_id": file_id,
                    }
                )
                self.state_repository.mark_interrupted(drive_folder_id, file_id)
                report.continuation_token = file_id
                break

            records, parsing_errors = self._parse_excel_rows(
                excel_bytes=file_bytes,
                excel_file_id=file_id,
            )

            if parsing_errors:
                report.errors.extend(parsing_errors)

            if not records:
                self.logger.info(
                    "ℹ️ Archivo sin registros nuevos",
                    etapa="parseo",
                    drive_folder_id=drive_folder_id,
                    excel_file_id=file_id,
                    excel_filename=filename,
                    table_rows=0,
                    message_id=None,
                )
                report.processed_files.append(
                    {
                        "file_id": file_id,
                        "filename": filename,
                        "inserted": 0,
                        "skipped": 0,
                        "rows_total": 0,
                    }
                )
                self.state_repository.mark_processed(drive_folder_id, file_id)
                continue

            try:
                inserted, skipped = self._persist_records(records)
            except SQLAlchemyError as exc:
                self.logger.error(
                    "❌ Error insertando registros en MySQL",
                    etapa="insercion_mysql",
                    drive_folder_id=drive_folder_id,
                    excel_file_id=file_id,
                    excel_filename=filename,
                    error=str(exc),
                    error_code="db_insert_failed",
                    table_errors=[],
                    message_id=None,
                )
                report.errors.append(
                    {
                        "code": "db_insert_failed",
                        "message": str(exc),
                        "file_id": file_id,
                    }
                )
                self.state_repository.mark_interrupted(drive_folder_id, file_id)
                report.continuation_token = file_id
                break

            inserted_total += inserted
            skipped_total += skipped

            self.logger.info(
                "✅ Archivo procesado",
                etapa="insercion_mysql",
                drive_folder_id=drive_folder_id,
                excel_file_id=file_id,
                excel_filename=filename,
                records_processed=inserted,
                records_skipped=skipped,
                table_rows=len(records),
                table_errors=[],
                message_id=None,
            )

            report.processed_files.append(
                {
                    "file_id": file_id,
                    "filename": filename,
                    "inserted": inserted,
                    "skipped": skipped,
                    "rows_total": len(records),
                }
            )
            self.state_repository.mark_processed(drive_folder_id, file_id)
        else:
            # Se completa el ciclo
            self.state_repository.clear(drive_folder_id)

        report.inserted_count = inserted_total
        report.skipped_count = skipped_total
        report.duration_seconds = (datetime.utcnow() - start_time).total_seconds()
        if not report.continuation_token:
            report.continuation_token = None

        self.logger.info(
            "🎉 Sincronización de generación finalizada",
            etapa="fin",
            drive_folder_id=drive_folder_id,
            fecha_generacion=fecha_generacion,
            inserted_count=inserted_total,
            skipped_count=skipped_total,
            duration_seconds=report.duration_seconds,
            continuation_token=report.continuation_token,
            message_id=None,
        )
        return report

    def _parse_excel_rows(self, *, excel_bytes: bytes, excel_file_id: str) -> (List[MissionaryRecord], List[Dict[str, Any]]):
        errors: List[Dict[str, Any]] = []
        workbook = None
        try:
            workbook = load_workbook(io.BytesIO(excel_bytes), read_only=True, data_only=True)
        except Exception as exc:  # noqa: BLE001
            errors.append(
                {
                    "code": "excel_read_failed",
                    "message": str(exc),
                    "file_id": excel_file_id,
                }
            )
            return [], errors

        try:
            worksheet = workbook.worksheets[0]
        except IndexError:
            if workbook is not None:
                workbook.close()
            return [], errors

        rows = list(worksheet.iter_rows(values_only=True))
        if len(rows) < 2:
            workbook.close()
            return [], errors

        records: List[MissionaryRecord] = []
        for index, row in enumerate(rows[1:], start=2):
            row_values = list(row)
            record = MissionaryRecord.from_row(
                row_values,
                logger=self.logger,
                excel_file_id=excel_file_id,
                row_index=index,
            )
            if record:
                records.append(record)

        workbook.close()
        return records, errors

    def _persist_records(self, records: List[MissionaryRecord]) -> (int, int):
        if not records:
            return 0, 0

        record_ids = [record.id for record in records]
        timestamp = datetime.utcnow()

        with self.session_factory() as session:
            existing_ids = self._fetch_existing_ids(session, record_ids)
            new_records = [r for r in records if r.id not in existing_ids]

            if not new_records:
                return 0, len(records)

            payload = [record.to_database_payload(timestamp=timestamp) for record in new_records]

            session.execute(insert(ccm_generaciones_table), payload)
            session.commit()
            return len(new_records), len(records) - len(new_records)

    def _fetch_existing_ids(self, session: Session, ids: Iterable[int]) -> set[int]:
        id_list = list(ids)
        if not id_list:
            return set()
        result = session.execute(select(ccm_generaciones_table.c.id).where(ccm_generaciones_table.c.id.in_(id_list)))
        return {row[0] for row in result}


__all__ = [
    "DatabaseSyncService",
    "DatabaseSyncReport",
    "DatabaseSyncStateRepository",
    "MissionaryRecord",
]
